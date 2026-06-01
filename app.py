"""
MJT 모바일 현황 대시보드 — Flask
Google Sheets 데이터를 모바일 브라우저로 조회 + 오늘 메뉴 등록
"""
import os, json, base64, time, threading, secrets, io, re, functools
from datetime import datetime, date, timedelta, timezone
from collections import defaultdict
from flask import Flask, render_template, jsonify, request, redirect, url_for, session, send_file

import gspread
from gspread.exceptions import WorksheetNotFound
from google.oauth2.service_account import Credentials


def _norm_date(s) -> str:
    """Google Sheets 다양한 날짜 포맷 → 'YYYY-MM-DD' 정규화.
    시리얼 숫자(46169), '5/29/2026', '2026. 5. 29.' 등 모두 처리."""
    if isinstance(s, (int, float)) and s > 1000:
        try:
            return (date(1899, 12, 30) + timedelta(days=int(s))).strftime('%Y-%m-%d')
        except Exception:
            return str(s)
    s = str(s).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m: return f'{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
    m = re.match(r'^(\d{4})[./]\s*(\d{1,2})[./]\s*(\d{1,2})', s)
    if m: return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    return s

app = Flask(__name__)

# ⚠ SECRET_KEY: Render 환경변수 'SECRET_KEY' 필수 등록
# 미등록 시 재배포마다 새로 생성됨 → 기존 세션 모두 무효화
# 환경변수가 없으면 경고 표시 (개발 환경만 fallback)
_secret = os.environ.get('SECRET_KEY')
if not _secret:
    print('[⚠ WARN] SECRET_KEY 환경변수 미설정 — 재배포 시 세션 만료됨!')
    _secret = secrets.token_hex(32)
app.secret_key = _secret

# 세션 쿠키 옵션 명시 (30일 유지)
app.config.update(
    PERMANENT_SESSION_LIFETIME=timedelta(days=30),
    SESSION_REFRESH_EACH_REQUEST=True,    # 매 요청마다 쿠키 갱신
    SESSION_COOKIE_NAME='mjt_session',
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=True,           # HTTPS 전용 (Render는 HTTPS)
)
app.permanent_session_lifetime = timedelta(days=30)

# 식당 메뉴 등록 비밀번호
MENU_PW = os.environ.get('MENU_PW', '3838')

# ── 모바일 접속 인증 (v1.2.2 추가) ────────────────────────────
# 회사 공통 코드 (분기마다 갱신 권장). Render Environment 에 등록.
MOBILE_AUTH_CODE = os.environ.get('MOBILE_AUTH_CODE', 'mj3838')

# IP 기반 로그인 시도 차단 (5회 실패 → 30분)
_login_attempts: dict = {}   # ip → [timestamps]
_login_lock = threading.Lock()

# 인증 면제 경로 (로그인/정적 자산 등)
_AUTH_EXEMPT_PREFIX = ('/login', '/logout', '/static', '/api/health',
                       '/favicon.ico')


def _ip():
    return request.headers.get('X-Forwarded-For',
                                request.remote_addr or 'unknown').split(',')[0].strip()


def _is_blocked(ip):
    with _login_lock:
        e = _login_attempts.get(ip)
        if not e: return False
        recent = [t for t in e if time.time() - t < 1800]
        _login_attempts[ip] = recent
        return len(recent) >= 5


def _record_failure(ip):
    with _login_lock:
        e = _login_attempts.setdefault(ip, [])
        e.append(time.time())
        _login_attempts[ip] = [t for t in e if time.time() - t < 1800]


def _clear_failure(ip):
    with _login_lock:
        _login_attempts.pop(ip, None)


def _verify_emp_active(emp_id: str):
    """사번이 employees.active=TRUE 여야 통과. 퇴사자 자동 차단.
    반환: dict(emp_id, name, dept, factory, phone) 또는 None."""
    if not emp_id:
        return None
    emp_id = emp_id.strip().upper()
    try:
        import db_pg
        if db_pg.is_available():
            r = db_pg.query_one(
                "SELECT emp_id, name, dept, factory, active, phone FROM employees "
                "WHERE UPPER(emp_id) = %s LIMIT 1", (emp_id,))
            if r and r.get('active'):
                return {'emp_id': r['emp_id'], 'name': r['name'],
                        'dept': r['dept'], 'factory': r['factory'],
                        'phone': r.get('phone') or ''}
            return None
    except Exception:
        pass
    return None


def _phone_last4(phone: str) -> str:
    """전화번호에서 숫자만 추출 후 끝 4자리 반환."""
    import re as _re
    digits = _re.sub(r'\D', '', phone or '')
    return digits[-4:] if len(digits) >= 4 else ''


def _log_access(emp_id, ip, path, success=True):
    """접속 로그 (best effort — 실패해도 진행)."""
    try:
        import db_pg
        if db_pg.is_available():
            with db_pg.cursor() as cur:
                cur.execute(
                    "INSERT INTO access_logs (emp_id, ip, path, user_agent, success) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (emp_id, ip, path,
                     request.headers.get('User-Agent', '')[:200], success))
    except Exception:
        pass


def require_auth(f):
    """모든 페이지 보호 데코레이터. 매 진입마다 active 재검증."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        emp_id = session.get('emp_id')
        if not emp_id:
            return redirect(url_for('login', next=request.path))
        # 매번 active 검증 (퇴사 즉시 차단)
        emp = _verify_emp_active(emp_id)
        if not emp:
            session.clear()
            return redirect(url_for('login', err='expired'))
        # 사원 정보 session에 최신화
        session['emp_name'] = emp['name']
        session['factory'] = emp.get('factory', '')
        return f(*args, **kwargs)
    return wrapper


@app.before_request
def _auth_check():
    """모든 요청 전에 인증 확인. 면제 경로는 통과."""
    path = request.path
    if any(path.startswith(p) for p in _AUTH_EXEMPT_PREFIX):
        return None
    emp_id = session.get('emp_id')
    if not emp_id:
        # API 호출은 401 JSON, 페이지는 redirect
        if path.startswith('/api/'):
            return jsonify({'ok': False, 'error': '인증 필요'}), 401
        return redirect(url_for('login', next=path))
    # active 재검증
    emp = _verify_emp_active(emp_id)
    if not emp:
        session.clear()
        if path.startswith('/api/'):
            return jsonify({'ok': False, 'error': '만료된 계정'}), 401
        return redirect(url_for('login', err='expired'))
    session['emp_name'] = emp['name']
    session['factory'] = emp.get('factory', '')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        nxt = request.args.get('next', '/')
        err = request.args.get('err')
        # 이미 로그인된 상태면 다음 페이지로
        if session.get('emp_id') and not err:
            return redirect(nxt)
        return render_template('login.html', err=err, next=nxt, version='1.4.0')

    ip = _ip()
    if _is_blocked(ip):
        return render_template('login.html', err='blocked', next='/')

    emp_id = request.form.get('emp_id', '').strip().upper()
    phone4 = request.form.get('phone4', '').strip()
    code   = request.form.get('code', '').strip()
    nxt    = request.form.get('next', '/') or '/'

    # ① 회사 공통 코드 검증
    if not MOBILE_AUTH_CODE or code != MOBILE_AUTH_CODE:
        _record_failure(ip)
        _log_access(emp_id or '?', ip, '/login', success=False)
        return render_template('login.html', err='wrong_code',
                               next=nxt, emp_id=emp_id)

    # ② 사번 + active 검증
    emp = _verify_emp_active(emp_id)
    if not emp:
        _record_failure(ip)
        _log_access(emp_id, ip, '/login', success=False)
        return render_template('login.html', err='not_emp',
                               next=nxt, emp_id=emp_id)

    # ③ 휴대폰 끝 4자리 검증
    db_last4 = _phone_last4(emp.get('phone', ''))
    if not db_last4:
        # DB에 전화번호 없는 사원 → 관리자에게 문의 안내
        _record_failure(ip)
        _log_access(emp_id, ip, '/login', success=False)
        return render_template('login.html', err='no_phone',
                               next=nxt, emp_id=emp_id)
    if phone4 != db_last4:
        _record_failure(ip)
        _log_access(emp_id, ip, '/login', success=False)
        return render_template('login.html', err='wrong_phone',
                               next=nxt, emp_id=emp_id)

    # 통과 → 세션 발급
    _clear_failure(ip)
    session.permanent = True
    session['emp_id']   = emp['emp_id']
    session['emp_name'] = emp['name']
    session['factory']  = emp.get('factory', '')
    _log_access(emp['emp_id'], ip, '/login', success=True)
    return redirect(nxt)


@app.route('/logout')
def logout():
    emp_id = session.get('emp_id')
    session.clear()
    if emp_id:
        _log_access(emp_id, _ip(), '/logout', success=True)
    return redirect(url_for('login'))


@app.route('/api/health')
def api_health():
    """헬스체크 (UptimeRobot 용 — 인증 면제)."""
    return jsonify({'ok': True, 'ts': datetime.now().isoformat()})

# ── 결재 설정 ────────────────────────────────────────────────────
# render.com 환경변수에 APPROVE_GRP_MJ_PW / APPROVE_CEO_MJ_PW /
# APPROVE_GRP_SCS_PW / APPROVE_DIR_SCS_PW 로 각 결재자 비밀번호 설정
APPROVAL_CFG = {
    'MJ': {
        'grp': {'name': '이광희', 'title': '제조그룹장', 'stamp': '이광희',
                'pw': os.environ.get('APPROVE_GRP_MJ_PW', '')},
        'ceo': {'name': '김신욱', 'title': '대표이사',   'stamp': '김신욱',
                'pw': os.environ.get('APPROVE_CEO_MJ_PW', '')},
    },
    'SCS': {
        'grp': {'name': '이현승', 'title': '제조팀장',   'stamp': '이현승',
                'pw': os.environ.get('APPROVE_GRP_SCS_PW', '')},
        'ceo': {'name': '김멋진', 'title': '이사',       'stamp': '김멋진',
                'pw': os.environ.get('APPROVE_DIR_SCS_PW', '')},
    },
}

# ── 상수 ────────────────────────────────────────────────────────
SHEET_NAME   = 'MJT_식수관리'
SCOPES       = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
ATT_CAT_OT   = ['잔업', '특근']
ATT_CAT_LEAVE= ['연차','오전반차','오후반차','탄력오전','탄력오후','계절휴가','하계휴가','기휴']
OT_DANGER_H  = 12   # 주 52H 한도 OT
OT_MAX64_H   = 24   # 주 64H 한도 OT
CACHE_TTL    = 300  # 5분 캐시

# ── 시간대 (v1.3 — 필리핀 등 다국가 대비) ────────────────────
# 사용자 시간대 우선순위: 쿠키 → 사번 prefix(M/S=KST, P=PHT) → 환경변수 → KST
KST = timezone(timedelta(hours=9))   # 한국
PHT = timezone(timedelta(hours=8))   # 필리핀
TZ_MAP = {'KST': KST, 'PHT': PHT, 'UTC': timezone(timedelta(hours=0))}
DEFAULT_TZ = os.environ.get('DEFAULT_TZ', 'KST')


def user_tz():
    """현재 요청의 시간대 결정.
    1) 쿠키 'tz' 값 (사용자가 명시 선택)
    2) 사번 prefix: P로 시작 → 필리핀, 그 외 → KST
    3) 환경변수 DEFAULT_TZ
    """
    try:
        from flask import session, request
        # 1) 쿠키 우선
        tz_name = request.cookies.get('tz')
        if tz_name and tz_name in TZ_MAP:
            return TZ_MAP[tz_name]
        # 2) 사번 prefix
        emp_id = session.get('emp_id', '') if session else ''
        if emp_id.upper().startswith('P'):
            return PHT
    except Exception:
        pass
    return TZ_MAP.get(DEFAULT_TZ, KST)


def fmt_local(dt, fmt='%Y-%m-%d %H:%M'):
    """DB의 UTC TIMESTAMPTZ → 사용자 현지 시간 표시."""
    if dt is None: return ''
    if isinstance(dt, str): return dt
    try:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone(timedelta(hours=0)))
        return dt.astimezone(user_tz()).strftime(fmt)
    except Exception:
        return str(dt)

_OP_DEFAULTS = {
    '중식신청마감': '08:10',
    '중식시작':     '11:00',
    '중식마감':     '13:30',
    '저녁마감':     '15:00',
    '식당유형':     '식당',      # '식당' 또는 '중국집'
}

DUTY_SHEET = '당직자'

# ── 인증 ────────────────────────────────────────────────────────
def _get_creds():
    env_val = os.environ.get('GOOGLE_CREDS_JSON')
    if env_val:
        info = json.loads(base64.b64decode(env_val).decode('utf-8'))
        return Credentials.from_service_account_info(info, scopes=SCOPES)
    cred_file = os.path.join(os.path.dirname(__file__), '..',
                             'youtubehotfinder-471114-52ceb99f60d8.json')
    return Credentials.from_service_account_file(cred_file, scopes=SCOPES)

def _open_sh():
    gc = gspread.authorize(_get_creds())
    try:
        gc.session.timeout = (15, 60)
    except Exception:
        pass
    return gc.open(SHEET_NAME)

# ── 캐시 ────────────────────────────────────────────────────────
_cache: dict = {}
_lock = threading.Lock()
_guest_cache: list = []   # 세션 중 등록된 외부손님 (Sheets 포맷 불일치 대응)

def _cached(key, fn, ttl=CACHE_TTL):
    with _lock:
        e = _cache.get(key)
        if e and time.time() - e['t'] < ttl:
            return e['d']
    d = fn()
    with _lock:
        _cache[key] = {'d': d, 't': time.time()}
    return d

def _clear_cache():
    with _lock:
        _cache.clear()

# ── 데이터 함수 ──────────────────────────────────────────────────
def get_employees():
    def _f():
        sh = _open_sh()
        rows = sh.worksheet('사원마스터').get_all_records()
        return [r for r in rows if str(r.get('사용여부', 'Y')).strip() == 'Y']
    return _cached('emps', _f)

def get_att_records(year, month):
    def _f():
        sh = _open_sh()
        rows = sh.worksheet('근태기록').get_all_records()
        return [r for r in rows
                if str(r.get('연도', '')).strip() == str(year)
                and str(r.get('월', '')).strip() == str(month)]
    return _cached(f'att_{year}_{month}', _f)

def _now_kst() -> str:
    return datetime.now(KST).strftime('%H:%M')

def get_op_settings() -> dict:
    def _f():
        result = dict(_OP_DEFAULTS)
        try:
            sh   = _open_sh()
            rows = sh.worksheet('운영설정').get_all_values()[1:]
            for r in rows:
                if len(r) >= 2 and r[0].strip() and r[1].strip():
                    result[r[0].strip()] = r[1].strip()
        except Exception:
            pass
        return result
    return _cached('op_settings', _f, ttl=300)


def get_today_menu(ds: str) -> str:
    def _f():
        try:
            sh = _open_sh()
            rows = sh.worksheet('오늘점심메뉴').get_all_values()[1:]
            for r in rows:
                if len(r) >= 2 and r[0].strip() == ds:
                    return r[1].strip()
        except Exception:
            pass
        return ''
    return _cached(f'menu_{ds}', _f, ttl=60)

def save_today_menu(ds: str, menu: str):
    sh = _open_sh()
    ws = sh.worksheet('오늘점심메뉴')
    rows = ws.get_all_values()
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for i, r in enumerate(rows):
        if i == 0:
            continue
        if r and r[0].strip() == ds:
            ws.update(range_name=f'A{i+1}:D{i+1}',
                      values=[[ds, menu, '식당', now_s]])
            with _lock:
                _cache.pop(f'menu_{ds}', None)
            return
    ws.append_row([ds, menu, '식당', now_s], value_input_option='USER_ENTERED')
    with _lock:
        _cache.pop(f'menu_{ds}', None)

def get_meal_today(ds: str) -> dict:
    # 중식신청:   [날짜, 시간, 공장, 사원번호, 성명, 부서, 상태]
    # 중식실식수: [날짜, 시간, 유형, 사원번호, 성명, 부서, 직급, 메모, 공장]
    # 저녁도시락: [날짜, 시간, 사원번호, 성명, 부서, 직급, 성별, 사유]
    # 특근식사:   [날짜, 시간, 사원번호, 성명, 부서, 직급, mode, 메뉴, ...]
    # 외부손님:   [방문날짜, 등록날짜, 시간, 담당자id, 담당자성명, 부서, 회사명, 손님성명, 사유]
    def _f():
        sh = _open_sh()
        out = {}
        def _get(name, date_col=0):
            try:
                rows = sh.worksheet(name).get_all_values()[1:]
                return [r for r in rows if len(r) > date_col and r[date_col].strip() == ds]
            except Exception:
                return []
        def _get_guest():
            """외부손님 — 날짜 시리얼/문자열 모두 정규화 후 비교 + 세션 캐시 병합."""
            try:
                raw = sh.worksheet('외부손님').get_all_values(
                    value_render_option='UNFORMATTED_VALUE')[1:]
            except TypeError:
                raw = sh.worksheet('외부손님').get_all_values()[1:]
            except Exception:
                raw = []
            matched, seen = [], set()
            for r in raw:
                if not r: continue
                d0 = _norm_date(r[0]) if len(r) > 0 else ''
                d1 = _norm_date(r[1]) if len(r) > 1 else ''
                if d0 == ds or d1 == ds:
                    norm = list(r)
                    if len(norm) > 0: norm[0] = d0
                    if len(norm) > 1: norm[1] = d1
                    matched.append(norm)
                    if len(r) > 2: seen.add((d1, str(r[2]).strip()))
            for r in _guest_cache:
                if not r: continue
                d0 = _norm_date(r[0]) if len(r) > 0 else ''
                d1 = _norm_date(r[1]) if len(r) > 1 else ''
                if (d0 == ds or d1 == ds) and (d1, str(r[2]).strip()) not in seen:
                    matched.append(r)
            return matched
        out['중식신청']   = _get('중식신청')
        out['중식실식수'] = _get('중식실식수')
        out['저녁도시락'] = _get('저녁도시락')
        out['특근식사']   = _get('특근식사')
        out['외부손님']   = _get_guest()
        return out
    return _cached(f'meal_{ds}', _f, ttl=120)  # 2분 캐시

# ── 52H 위험도 계산 ──────────────────────────────────────────────
def calc_ot_status(emps, att_rows, year, month):
    fw = date(year, month, 1).weekday()
    def wn(d):
        return max(1, min(6, (d + fw - 1) // 7 + 1))

    emp_map = {str(e.get('사원번호', '')).strip(): e for e in emps}
    emp_wk: dict = {}

    for r in att_rows:
        eid   = str(r.get('사원번호', '')).strip()
        atype = str(r.get('근태유형', '')).strip()
        try:   v = float(r.get('값', 0) or 0)
        except: v = 0.0
        try:   day = int(str(r.get('일자', '')).split('-')[-1])
        except: continue
        w = wn(day)
        if eid not in emp_wk:
            emp_wk[eid] = {'ot': {i: 0.0 for i in range(1, 7)},
                           'lv': {i: 0.0 for i in range(1, 7)},
                           'jn': 0.0, 'sp': 0.0}
        if atype == '잔업':
            emp_wk[eid]['ot'][w] += v; emp_wk[eid]['jn'] += v
        elif atype == '특근':
            emp_wk[eid]['ot'][w] += v; emp_wk[eid]['sp'] += v
        elif atype in ATT_CAT_LEAVE:
            emp_wk[eid]['lv'][w] += v * 8

    rows = []
    for eid, wk in emp_wk.items():
        e = emp_map.get(eid, {})
        avail52  = {w: OT_DANGER_H + wk['lv'][w] for w in range(1, 7)}
        avail64  = {w: OT_MAX64_H  + wk['lv'][w] for w in range(1, 7)}
        worst52  = max(wk['ot'][w] - avail52[w] for w in range(1, 7))
        worst64  = max(wk['ot'][w] - avail64[w] for w in range(1, 7))
        worst_wk = max(wk['ot'].values())
        total_ot = wk['jn'] + wk['sp']
        if   worst64 >= 0:    status = 'max64'
        elif worst52 >= 0:    status = 'danger'
        elif worst52 >= -4:   status = 'warn'
        else:                 status = 'safe'
        rows.append({
            'eid': eid,
            'name': e.get('성명', eid),
            'dept': e.get('부서명', '?'),
            'total_ot': total_ot,
            'jn':  wk['jn'],
            'sp':  wk['sp'],
            'worst_wk': worst_wk,
            'status': status,
        })
    status_order = {'max64': 0, 'danger': 1, 'warn': 2, 'safe': 3}
    rows.sort(key=lambda r: (status_order[r['status']], -r['total_ot']))
    return rows

# ── 공휴일 DB ────────────────────────────────────────────────────
def _build_holiday_db():
    db = {}
    fixed = {
        '01-01': '신정',      '03-01': '삼일절',     '05-01': '근로자의날',
        '05-05': '어린이날',  '06-06': '현충일',     '08-15': '광복절',
        '10-03': '개천절',    '10-09': '한글날',     '12-25': '성탄절',
    }
    for yr in range(2025, 2051):
        for mmdd, name in fixed.items():
            db[f'{yr}-{mmdd}'] = name
    lunar = {
        '2026-02-16':'설날 연휴','2026-02-17':'설날','2026-02-18':'설날 연휴',
        '2026-05-25':'부처님오신날',
        '2026-09-24':'추석 연휴','2026-09-25':'추석','2026-09-26':'추석 연휴',
        '2027-02-05':'설날 연휴','2027-02-06':'설날','2027-02-07':'설날 연휴',
        '2027-05-13':'부처님오신날',
        '2027-09-14':'추석 연휴','2027-09-15':'추석','2027-09-16':'추석 연휴',
        '2028-01-26':'설날 연휴','2028-01-27':'설날','2028-01-28':'설날 연휴',
        '2028-05-02':'부처님오신날',
        '2028-10-02':'추석 연휴','2028-10-03':'추석','2028-10-04':'추석 연휴',
        '2029-02-12':'설날 연휴','2029-02-13':'설날','2029-02-14':'설날 연휴',
        '2029-05-21':'부처님오신날',
        '2029-10-05':'추석','2029-10-06':'추석 연휴',
        '2030-02-02':'설날 연휴','2030-02-03':'설날','2030-02-04':'설날 연휴',
        '2030-05-11':'부처님오신날',
        '2030-09-22':'추석 연휴','2030-09-23':'추석','2030-09-24':'추석 연휴',
        '2031-01-22':'설날 연휴','2031-01-23':'설날','2031-01-24':'설날 연휴',
        '2031-05-28':'부처님오신날',
        '2031-09-11':'추석 연휴','2031-09-12':'추석','2031-09-13':'추석 연휴',
        '2032-02-10':'설날 연휴','2032-02-11':'설날','2032-02-12':'설날 연휴',
        '2032-05-16':'부처님오신날',
        '2032-09-29':'추석 연휴','2032-09-30':'추석','2032-10-01':'추석 연휴',
        '2033-01-30':'설날 연휴','2033-01-31':'설날','2033-02-01':'설날 연휴',
        '2033-05-05':'부처님오신날',
        '2033-09-18':'추석 연휴','2033-09-19':'추석','2033-09-20':'추석 연휴',
        '2034-02-18':'설날 연휴','2034-02-19':'설날','2034-02-20':'설날 연휴',
        '2034-05-25':'부처님오신날',
        '2034-10-07':'추석 연휴','2034-10-08':'추석','2034-10-09':'추석 연휴',
        '2035-02-07':'설날 연휴','2035-02-08':'설날','2035-02-09':'설날 연휴',
        '2035-05-14':'부처님오신날',
        '2035-09-26':'추석 연휴','2035-09-27':'추석','2035-09-28':'추석 연휴',
    }
    db.update(lunar)
    return db

BASE_HOLIDAYS = _build_holiday_db()

def get_managed_holidays():
    def _f():
        try:
            sh = _open_sh()
            rows = sh.worksheet('공휴일').get_all_values()[1:]
            return {r[0].strip(): r[1].strip()
                    for r in rows if len(r) >= 2 and r[0].strip()}
        except Exception:
            return {}
    return _cached('managed_hols', _f, ttl=3600)

def get_day_label(ds: str):
    """Returns '' for workday, '토'/'일'/holiday_name for non-workday."""
    try:
        d = date.fromisoformat(ds)
    except Exception:
        return ''
    wd = d.weekday()
    if wd == 5: return '토'
    if wd == 6: return '일'
    name = BASE_HOLIDAYS.get(ds) or get_managed_holidays().get(ds, '')
    return name

def _week_range(ref_ds=None):
    ref = date.fromisoformat(ref_ds) if ref_ds else date.today()
    mon = ref - timedelta(days=ref.weekday())
    return mon, mon + timedelta(days=6)

def get_week_att(ref_ds=None):
    mon, sun = _week_range(ref_ds)
    months = set()
    d = mon
    while d <= sun:
        months.add((d.year, d.month))
        d += timedelta(days=1)
    all_recs = []
    for yr, mo in months:
        all_recs.extend(get_att_records(yr, mo))
    week_dates = set()
    d = mon
    while d <= sun:
        week_dates.add(d.strftime('%Y-%m-%d'))
        d += timedelta(days=1)
    return [r for r in all_recs if str(r.get('일자', '')).strip() in week_dates], mon, sun

def _ot_time(hours):
    """Convert OT hours float to HH:MM start/end (assumes 08:00 start)."""
    try:
        h = float(hours)
    except Exception:
        h = 0.0
    start_min = 8 * 60
    work_min  = int(h * 60)
    lunch_min = 60 if h > 4 else 0
    end_min   = start_min + work_min + lunch_min
    def fmt(m):
        return f'{m // 60:02d}:{m % 60:02d}'
    return fmt(start_min), fmt(end_min)

# ── 라우트 ──────────────────────────────────────────────────────
@app.route('/')
def index():
    today = date.today()
    ds    = today.strftime('%Y-%m-%d')
    year, month = today.year, today.month
    try:
        emps     = get_employees()
        att      = get_att_records(year, month)
        meal     = get_meal_today(ds)
        ot_rows  = calc_ot_status(emps, att, year, month)
        n_warn   = sum(1 for r in ot_rows if r['status'] == 'warn')
        n_danger = sum(1 for r in ot_rows if r['status'] == 'danger')
        n_max64  = sum(1 for r in ot_rows if r['status'] == 'max64')
        total_ot = sum(r['total_ot'] for r in ot_rows)

        lunch_req  = meal['중식신청']
        lunch_real = [r for r in meal['중식실식수']
                      if len(r) > 2 and r[2].strip() == '중식']
        dinner     = meal['저녁도시락']
        wkend      = meal['특근식사']
        guests     = meal['외부손님']
        today_menu = get_today_menu(ds)

        return render_template('index.html',
            today   = today.strftime('%Y년 %m월 %d일'),
            weekday = '월화수목금토일'[today.weekday()],
            year=year, month=month,
            n_emps   = len(emps),
            total_ot = int(total_ot),
            n_warn=n_warn, n_danger=n_danger, n_max64=n_max64,
            today_menu   = today_menu,
            n_lunch_req  = len(lunch_req),
            n_lunch_real = len(lunch_real),
            n_dinner     = len(dinner),
            n_wkend      = len(wkend),
            n_guest      = len(guests),
            updated = datetime.now().strftime('%H:%M'),
        )
    except Exception as e:
        return render_template('error.html', error=str(e))


@app.route('/overtime')
def overtime():
    today   = date.today()
    year    = int(request.args.get('year',  today.year))
    month   = int(request.args.get('month', today.month))
    factory = request.args.get('fac', '전체')
    try:
        emps = get_employees()
        if factory == '1공장':
            emps = [e for e in emps if str(e.get('사원번호','')).startswith('M')]
        elif factory == '2공장':
            emps = [e for e in emps if str(e.get('사원번호','')).startswith('S')]

        att = get_att_records(year, month)
        if factory != '전체':
            eids = {str(e.get('사원번호','')).strip() for e in emps}
            att  = [r for r in att if str(r.get('사원번호','')).strip() in eids]

        ot_rows  = calc_ot_status(emps, att, year, month)
        n_max64  = sum(1 for r in ot_rows if r['status'] == 'max64')
        n_danger = sum(1 for r in ot_rows if r['status'] == 'danger')
        n_warn   = sum(1 for r in ot_rows if r['status'] == 'warn')
        n_safe   = sum(1 for r in ot_rows if r['status'] == 'safe')

        # 이전/다음 월 계산
        prev_m = month - 1 if month > 1 else 12
        prev_y = year if month > 1 else year - 1
        next_m = month + 1 if month < 12 else 1
        next_y = year if month < 12 else year + 1

        return render_template('overtime.html',
            ot_rows=ot_rows, year=year, month=month, factory=factory,
            n_total=len(emps),
            n_max64=n_max64, n_danger=n_danger, n_warn=n_warn, n_safe=n_safe,
            prev_y=prev_y, prev_m=prev_m,
            next_y=next_y, next_m=next_m,
            updated=datetime.now().strftime('%H:%M'),
        )
    except Exception as e:
        return render_template('error.html', error=str(e))


@app.route('/meal')
def meal():
    today = date.today()
    ds    = request.args.get('date', today.strftime('%Y-%m-%d'))
    try:
        meal_data = get_meal_today(ds)

        lunch_req  = meal_data['중식신청']
        lunch_real = [r for r in meal_data['중식실식수']
                      if len(r) > 2 and r[2].strip() == '중식']
        dinner_m   = [r for r in meal_data['저녁도시락']
                      if len(r) <= 6 or r[6].strip() != '여']
        dinner_f   = [r for r in meal_data['저녁도시락']
                      if len(r) > 6 and r[6].strip() == '여']
        wkend_g    = [r for r in meal_data['특근식사']
                      if len(r) > 6 and r[6].strip() == '구내식당']
        wkend_c    = [r for r in meal_data['특근식사']
                      if len(r) > 6 and r[6].strip() == '중국집']
        guests     = meal_data['외부손님']

        today_menu = get_today_menu(ds)
        return render_template('meal.html',
            ds=ds,
            today_menu=today_menu,
            lunch_req=lunch_req, lunch_real=lunch_real,
            dinner_m=dinner_m, dinner_f=dinner_f,
            wkend_g=wkend_g, wkend_c=wkend_c,
            guests=guests,
            updated=datetime.now().strftime('%H:%M'),
        )
    except Exception as e:
        return render_template('error.html', error=str(e))


@app.route('/menu', methods=['GET', 'POST'])
def menu_edit():
    today = date.today()
    ds    = today.strftime('%Y-%m-%d')
    error = ''
    saved = False

    # 로그인 처리
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'login':
            if request.form.get('pw') == MENU_PW:
                session['menu_auth'] = True
            else:
                error = '비밀번호가 틀렸습니다.'
        elif action == 'save' and session.get('menu_auth'):
            menu_text = request.form.get('menu', '').strip()
            if menu_text:
                try:
                    save_today_menu(ds, menu_text)
                    saved = True
                except Exception as e:
                    error = str(e)
        elif action == 'logout':
            session.pop('menu_auth', None)
            return redirect(url_for('menu_edit'))

    current_menu = get_today_menu(ds) if session.get('menu_auth') else ''
    return render_template('menu_edit.html',
        ds=ds,
        authed=session.get('menu_auth', False),
        current_menu=current_menu,
        error=error,
        saved=saved,
        weekday='월화수목금토일'[today.weekday()],
    )


@app.route('/ot_schedule')
def ot_schedule():
    today  = date.today()
    ds     = today.strftime('%Y-%m-%d')
    ref_ds = request.args.get('week', ds)

    try:
        # ── 금일 잔업/특근 ─────────────────────────────
        today_recs = get_att_records(today.year, today.month)
        today_jn = [r for r in today_recs
                    if str(r.get('일자', '')).strip() == ds
                    and r.get('근태유형', '') == '잔업']
        today_sp = [r for r in today_recs
                    if str(r.get('일자', '')).strip() == ds
                    and r.get('근태유형', '') == '특근']

        # ── 금주 휴일 특근 일정 ─────────────────────────
        week_recs, mon, sun = get_week_att(ref_ds)

        # 특근 중 휴일(토/일/공휴일)인 날만
        hol_sp = []
        for r in week_recs:
            if r.get('근태유형', '') != '특근':
                continue
            ds_r  = str(r.get('일자', '')).strip()
            label = get_day_label(ds_r)
            if label:
                hol_sp.append(dict(r, _label=label))

        # 날짜별 그룹핑
        grouped = defaultdict(list)
        for r in sorted(hol_sp, key=lambda x: str(x.get('일자', ''))):
            grouped[str(r.get('일자', ''))].append(r)

        # 주간 사원별 총OT (주40H 초과시간 계산용)
        emp_week_ot = defaultdict(float)
        for r in week_recs:
            if r.get('근태유형', '') in ATT_CAT_OT:
                try:
                    emp_week_ot[str(r.get('사원번호', '')).strip()] += float(r.get('값', 0) or 0)
                except Exception:
                    pass

        # 이번 주 토/일 당직자
        op        = get_op_settings()
        menu_type = op.get('식당유형', '식당')
        duty_map  = {}
        for delta in range(7):
            chk = mon + timedelta(days=delta)
            if chk.weekday() in (5, 6):
                cs = chk.strftime('%Y-%m-%d')
                duty_map[cs] = get_duty_for_date(cs) or '미배정'

        return render_template('ot_schedule.html',
            today_ds      = ds,
            today_weekday = '월화수목금토일'[today.weekday()],
            today_jn      = today_jn,
            today_sp      = today_sp,
            grouped_sp    = dict(grouped),
            mon = mon.strftime('%Y-%m-%d'),
            sun = sun.strftime('%Y-%m-%d'),
            ref_ds        = ref_ds,
            emp_week_ot   = dict(emp_week_ot),
            duty_map      = duty_map,
            menu_type     = menu_type,
            updated       = datetime.now().strftime('%H:%M'),
        )
    except Exception as e:
        return render_template('error.html', error=str(e))


def _ot_week_data(ref_ds):
    """미리보기/내보내기 공통 데이터. (hol_sp, emp_week_ot, mon, sun) 반환."""
    week_recs, mon, sun = get_week_att(ref_ds)
    hol_sp = []
    for r in week_recs:
        if r.get('근태유형', '') != '특근':
            continue
        ds_r  = str(r.get('일자', '')).strip()
        label = get_day_label(ds_r)
        if label:
            hol_sp.append(dict(r, _label=label))
    hol_sp.sort(key=lambda x: str(x.get('일자', '')))
    emp_week_ot = defaultdict(float)
    for r in week_recs:
        if r.get('근태유형', '') in ATT_CAT_OT:
            try:
                emp_week_ot[str(r.get('사원번호', '')).strip()] += float(r.get('값', 0) or 0)
            except Exception:
                pass
    return hol_sp, dict(emp_week_ot), mon, sun


@app.route('/ot_schedule/preview')
def ot_schedule_preview():
    ref_ds  = request.args.get('week', date.today().strftime('%Y-%m-%d'))
    fac     = request.args.get('fac', 'MJ')   # 'MJ' or 'SCS'
    appr_err= request.args.get('appr_err', '')
    appr_ok = request.args.get('appr_ok', '')
    if fac not in ('MJ', 'SCS'):
        fac = 'MJ'
    fac_prefix = 'M' if fac == 'MJ' else 'S'

    try:
        hol_sp, emp_week_ot, mon, sun = _ot_week_data(ref_ds)

        # 공장별 필터
        hol_sp = [r for r in hol_sp
                  if str(r.get('사원번호', '')).strip().upper().startswith(fac_prefix.upper())]

        # 날짜 목록 (unique, sorted)
        unique_dates = sorted(set(str(r.get('일자', '')).strip() for r in hol_sp))

        # GL 이름 — query param gl_YYYYMMDD=이름
        date_labels = []
        for ds in unique_dates:
            gl = request.args.get(f'gl_{ds.replace("-","")}', '')
            try:
                d_obj = date.fromisoformat(ds)
                label = f'{d_obj.strftime("%m/%d")} ({get_day_label(ds)})'
            except Exception:
                label = ds
            date_labels.append({'ds': ds, 'label': label, 'gl': gl})

        # 사람별 피벗
        person_map: dict = {}
        for r in hol_sp:
            ds_r  = str(r.get('일자', '')).strip()
            eid   = str(r.get('사원번호', '')).strip()
            value = r.get('값', 0)
            ts, te = _ot_time(value)
            note   = str(r.get('비고', '')).strip()
            if eid not in person_map:
                person_map[eid] = {
                    'dept': str(r.get('부서명', '')).strip(),
                    'name': str(r.get('성명', '')).strip(),
                    'task': note,
                    'days': {},
                }
            person_map[eid]['days'][ds_r] = f'{ts}~{te}'
            if note and note not in person_map[eid]['task']:
                sep = ', ' if person_map[eid]['task'] else ''
                person_map[eid]['task'] += sep + note

        # 부서 순서 (등장 순서 기준)
        dept_order: list = []
        for p in person_map.values():
            if p['dept'] not in dept_order:
                dept_order.append(p['dept'])

        persons = sorted(person_map.items(),
                         key=lambda x: (dept_order.index(x[1]['dept'])
                                        if x[1]['dept'] in dept_order else 999,
                                        x[1]['name']))

        # rowspan 계산 (dept 연속 그룹)
        rows = []
        no = 1
        i = 0
        while i < len(persons):
            dept = persons[i][1]['dept']
            j = i + 1
            while j < len(persons) and persons[j][1]['dept'] == dept:
                j += 1
            span = j - i
            for k in range(i, j):
                _, p = persons[k]
                rows.append({
                    'no':        no,
                    'dept':      dept,
                    'dept_span': span if k == i else 0,
                    'name':      p['name'],
                    'task':      p['task'],
                    'days':      p['days'],
                })
                no += 1
            i = j

        # 날짜별 인원 합계
        totals = {ds: sum(1 for _, p in persons if ds in p['days'])
                  for ds in unique_dates}

        # title 생성
        if unique_dates:
            d0 = date.fromisoformat(unique_dates[0])
            d1 = date.fromisoformat(unique_dates[-1])
            period = f'{d0.strftime("%m/%d")}~{d1.strftime("%m/%d")}'
        else:
            period = f'{mon.strftime("%m/%d")}~{sun.strftime("%m/%d")}'
        doc_title = f'{mon.strftime("%m")}월 특근 근무자 명단 ({period})'

        # 결재 상태
        week_key = mon.strftime('%Y-%m-%d')
        appr     = get_approval(week_key, fac)
        raw_cfg  = APPROVAL_CFG.get(fac, {})
        # 비밀번호 제외하고 템플릿에 전달
        appr_cfg = {k: {ik: iv for ik, iv in v.items() if ik != 'pw'}
                    for k, v in raw_cfg.items()}

        # 도장 이미지 base64 인코딩 (없으면 None)
        def _stamp_b64(name):
            try:
                p = os.path.join(os.path.dirname(__file__), 'static', 'stamps', f'{name}.png')
                with open(p, 'rb') as f:
                    return base64.b64encode(f.read()).decode()
            except Exception:
                return None

        grp_stamp = _stamp_b64(appr_cfg.get('grp', {}).get('stamp', '')) if appr_cfg else None
        ceo_stamp = _stamp_b64(appr_cfg.get('ceo', {}).get('stamp', '')) if appr_cfg else None

        # 당직자 + 식당유형
        op_cfg    = get_op_settings()
        menu_type = op_cfg.get('식당유형', '식당')
        duty_list = []
        for delta in range(7):
            chk = mon + timedelta(days=delta)
            if chk.weekday() in (5, 6):
                cs   = chk.strftime('%Y-%m-%d')
                name = get_duty_for_date(cs) or '미배정'
                duty_list.append({'date': cs, 'weekday': '토' if chk.weekday()==5 else '일', 'name': name})

        return render_template('ot_preview.html',
            rows        = rows,
            date_labels = date_labels,
            unique_dates= unique_dates,
            totals      = totals,
            doc_title   = doc_title,
            mon         = mon,
            sun         = sun,
            ref_ds      = ref_ds,
            fac         = fac,
            week_key    = week_key,
            appr        = appr,
            appr_cfg    = appr_cfg,
            grp_stamp   = grp_stamp,
            ceo_stamp   = ceo_stamp,
            appr_err    = appr_err,
            appr_ok     = appr_ok,
            today_str   = date.today().strftime('%Y-%m-%d'),
            menu_type   = menu_type,
            duty_list   = duty_list,
        )
    except Exception as e:
        return render_template('error.html', error=str(e))


@app.route('/ot_schedule/export')
def ot_schedule_export():
    ref_ds = request.args.get('week', date.today().strftime('%Y-%m-%d'))
    fac    = request.args.get('fac', 'MJ')
    if fac not in ('MJ', 'SCS'):
        fac = 'MJ'
    fac_prefix = 'M' if fac == 'MJ' else 'S'
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter

        hol_sp, emp_week_ot, mon, sun = _ot_week_data(ref_ds)
        hol_sp = [r for r in hol_sp
                  if str(r.get('사원번호', '')).strip().upper().startswith(fac_prefix.upper())]

        wb = Workbook()
        ws = wb.active
        ws.title = '특근신청서'

        # 컬럼 너비
        col_w = [6, 16, 10, 10, 14, 10, 30, 14]
        for i, w in enumerate(col_w, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 스타일 헬퍼
        def bd(style='thin'):
            s = Side(style=style)
            return Border(left=s, right=s, top=s, bottom=s)
        def fill(hex_color):
            return PatternFill('solid', fgColor=hex_color)
        def aln(h='center', v='center', wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        # ── 타이틀 ─────────────────────────────────────────
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 32
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 8

        ws.merge_cells('A2:H2')
        tc = ws['A2']
        tc.value     = 'MJT 주식회사  휴일 특근 신청서'
        tc.font      = Font(size=18, bold=True, color='1A237E')
        tc.alignment = aln('center')

        ws.merge_cells('A3:H3')
        dc = ws['A3']
        period_str = f'{mon.strftime("%Y년 %m월 %d일")} ~ {sun.strftime("%m월 %d일")}  |  작성일 {date.today().strftime("%Y-%m-%d")}'
        dc.value     = period_str
        dc.font      = Font(size=10, color='555555')
        dc.alignment = aln('center')

        # ── 헤더 행 ────────────────────────────────────────
        HDR_ROW = 5
        ws.row_dimensions[HDR_ROW].height = 22
        headers = ['No', '근무일', '시작시간', '종료시간', '공정(부서)', '성명', '근로 사유', '주40H 초과시간']
        hdr_fill = fill('1565C0')
        for col, h in enumerate(headers, 1):
            c = ws.cell(HDR_ROW, col, h)
            c.font      = Font(bold=True, color='FFFFFF', size=10)
            c.fill      = hdr_fill
            c.alignment = aln('center')
            c.border    = bd()

        # ── 데이터 행 ───────────────────────────────────────
        DATE_COLORS = ['E8F5E9', 'E3F2FD', 'FFF3E0', 'F3E5F5', 'FCE4EC', 'E0F7FA']
        date_color_map = {}
        color_idx = 0

        row_num = HDR_ROW + 1
        no = 1
        for r in hol_sp:
            ds_r   = str(r.get('일자', '')).strip()
            label  = r.get('_label', '')
            dept   = str(r.get('부서명', '')).strip()
            name   = str(r.get('성명', '')).strip()
            value  = r.get('값', 0)
            note   = str(r.get('비고', '')).strip()
            eid    = str(r.get('사원번호', '')).strip()
            week_h = emp_week_ot.get(eid, float(value or 0))

            try:
                d_obj = date.fromisoformat(ds_r)
                day_str = f'{d_obj.strftime("%m/%d")} ({label})'
            except Exception:
                day_str = ds_r

            t_start, t_end = _ot_time(value)

            if ds_r not in date_color_map:
                date_color_map[ds_r] = DATE_COLORS[color_idx % len(DATE_COLORS)]
                color_idx += 1
            row_fill = fill(date_color_map[ds_r])

            ws.row_dimensions[row_num].height = 18
            row_data = [no, day_str, t_start, t_end, dept, name, note,
                        f'{week_h:.0f}H']
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row_num, col, val)
                c.fill      = row_fill
                c.border    = bd()
                c.font      = Font(size=10)
                c.alignment = aln('center' if col != 7 else 'left', wrap=True)
            row_num += 1
            no += 1

        if no == 1:
            ws.merge_cells(f'A{row_num}:H{row_num}')
            c = ws.cell(row_num, 1, '해당 기간 휴일 특근 기록 없음')
            c.alignment = aln('center')
            c.font      = Font(italic=True, color='888888')
            row_num += 1

        # ── 결재 행 ────────────────────────────────────────
        row_num += 1
        ws.row_dimensions[row_num].height = 18
        ws.merge_cells(f'A{row_num}:B{row_num}')
        ws.cell(row_num, 1, '담당').font = Font(bold=True, size=10)
        ws.cell(row_num, 1).alignment    = aln('center')
        ws.cell(row_num, 1).border       = bd()
        ws.merge_cells(f'C{row_num}:D{row_num}')
        ws.cell(row_num, 3, '팀장').font  = Font(bold=True, size=10)
        ws.cell(row_num, 3).alignment     = aln('center')
        ws.cell(row_num, 3).border        = bd()
        ws.merge_cells(f'E{row_num}:F{row_num}')
        ws.cell(row_num, 5, '부문장').font = Font(bold=True, size=10)
        ws.cell(row_num, 5).alignment      = aln('center')
        ws.cell(row_num, 5).border         = bd()
        ws.merge_cells(f'G{row_num}:H{row_num}')
        ws.cell(row_num, 7, '대표이사').font = Font(bold=True, size=10)
        ws.cell(row_num, 7).alignment        = aln('center')
        ws.cell(row_num, 7).border           = bd()
        row_num += 1
        ws.row_dimensions[row_num].height = 40
        for col in [1, 3, 5, 7]:
            end = col + 1
            ws.merge_cells(f'{get_column_letter(col)}{row_num}:{get_column_letter(end)}{row_num}')
            ws.cell(row_num, col, '').border = bd()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'휴일특근신청서_{mon.strftime("%Y%m%d")}.xlsx'
        return send_file(buf, download_name=fname,
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return render_template('error.html', error=str(e))


# ── 결재 함수 ────────────────────────────────────────────────────
_APPR_SHEET = '특근결재'
# 컬럼: 주차키|공장|상태|1차자|1차시|1차결|1차사유|최종자|최종시|최종결|최종사유

def get_approval(week_key: str, factory: str) -> dict:
    def _f():
        try:
            sh   = _open_sh()
            rows = sh.worksheet(_APPR_SHEET).get_all_values()[1:]
            for r in rows:
                if len(r) >= 2 and r[0].strip() == week_key and r[1].strip() == factory:
                    def _g(i): return r[i].strip() if len(r) > i else ''
                    return {'week_key': _g(0), 'factory': _g(1), 'status': _g(2) or '대기',
                            'grp_name': _g(3), 'grp_dt': _g(4), 'grp_dec': _g(5), 'grp_reason': _g(6),
                            'ceo_name': _g(7), 'ceo_dt': _g(8), 'ceo_dec': _g(9), 'ceo_reason': _g(10)}
        except Exception:
            pass
        return {'week_key': week_key, 'factory': factory, 'status': '대기',
                'grp_name':'','grp_dt':'','grp_dec':'','grp_reason':'',
                'ceo_name':'','ceo_dt':'','ceo_dec':'','ceo_reason':''}
    return _cached(f'appr_{week_key}_{factory}', _f, ttl=60)

def _invalidate_approval(week_key, factory):
    with _lock:
        _cache.pop(f'appr_{week_key}_{factory}', None)

def save_approval(week_key, factory, level, decision, reason, approver_name):
    sh   = _open_sh()
    ws   = sh.worksheet(_APPR_SHEET)
    now  = datetime.now().strftime('%Y-%m-%d %H:%M')
    rows = ws.get_all_values()
    row_idx = None
    for i, r in enumerate(rows[1:], start=2):
        if len(r) >= 2 and r[0].strip() == week_key and r[1].strip() == factory:
            row_idx = i; break

    if level == 'grp':
        new_status = '1차완료' if decision == '승인' else '반려'
        if row_idx:
            ws.update(range_name=f'C{row_idx}:G{row_idx}',
                      values=[[new_status, approver_name, now, decision, reason]])
        else:
            ws.append_row([week_key, factory, new_status,
                           approver_name, now, decision, reason, '', '', '', ''],
                          value_input_option='USER_ENTERED')
    else:  # ceo
        new_status = '최종승인' if decision == '승인' else '반려'
        if row_idx:
            ws.update(range_name=f'C{row_idx}', values=[[new_status]])
            ws.update(range_name=f'H{row_idx}:K{row_idx}',
                      values=[[approver_name, now, decision, reason]])
        else:
            ws.append_row([week_key, factory, new_status,
                           '', '', '', '', approver_name, now, decision, reason],
                          value_input_option='USER_ENTERED')
    _invalidate_approval(week_key, factory)


@app.route('/ot_schedule/approve', methods=['POST'])
def ot_approve():
    week_key = request.form.get('week_key', '')
    factory  = request.form.get('factory', 'MJ')
    level    = request.form.get('level', 'grp')
    decision = request.form.get('decision', '승인')
    pw       = request.form.get('pw', '')
    reason   = request.form.get('reason', '').strip()
    ref_ds   = request.form.get('ref_ds', week_key)
    fac      = request.form.get('fac', factory)

    # 돌아갈 URL 구성
    back = f'/ot_schedule/preview?week={ref_ds}&fac={fac}'
    for key, val in request.form.items():
        if key.startswith('gl_'):
            back += f'&{key}={val}'

    if factory not in APPROVAL_CFG or level not in APPROVAL_CFG[factory]:
        return redirect(back + '&appr_err=설정오류')

    correct_pw = APPROVAL_CFG[factory][level]['pw']
    if not correct_pw:
        return redirect(back + '&appr_err=비밀번호미설정')
    if pw != correct_pw:
        return redirect(back + '&appr_err=비밀번호오류')

    state = get_approval(week_key, factory)
    status = state['status']
    if level == 'grp' and status not in ('대기', '반려'):
        return redirect(back + '&appr_err=이미1차처리됨')
    if level == 'ceo' and status != '1차완료':
        return redirect(back + '&appr_err=1차결재필요')

    name = APPROVAL_CFG[factory][level]['name']
    try:
        save_approval(week_key, factory, level, decision, reason, name)
    except Exception as e:
        return redirect(back + f'&appr_err={str(e)[:30]}')

    return redirect(back + '&appr_ok=1')


def get_company_events(year, month):
    def _f():
        try:
            sh   = _open_sh()
            rows = sh.worksheet('회사일정').get_all_values()[1:]
            result = []
            for r in rows:
                if len(r) < 3 or not r[0].strip():
                    continue
                ds = r[0].strip()
                if not (ds.startswith(f'{year}-{month:02d}') or
                        ds.startswith(f'{year}-{str(month).zfill(2)}')):
                    continue
                result.append({'ds': ds, 'type': r[1].strip(),
                                'content': r[2].strip(),
                                'note': r[3].strip() if len(r) > 3 else ''})
            return result
        except Exception:
            return []
    return _cached(f'events_{year}_{month}', _f, ttl=300)

def save_company_event(ds, ev_type, content, note=''):
    sh = _open_sh()
    ws = sh.worksheet('회사일정')
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append_row([ds, ev_type, content, note, now_s, '웹'], value_input_option='USER_ENTERED')
    with _lock:
        try:
            yr, mo = int(ds[:4]), int(ds[5:7])
            _cache.pop(f'events_{yr}_{mo}', None)
        except Exception:
            pass

def delete_company_event(ds, content):
    sh  = _open_sh()
    ws  = sh.worksheet('회사일정')
    all = ws.get_all_values()
    for i, row in enumerate(all[1:], start=2):
        if len(row) >= 3 and row[0].strip() == ds and row[2].strip() == content:
            ws.delete_rows(i)
            with _lock:
                try:
                    yr, mo = int(ds[:4]), int(ds[5:7])
                    _cache.pop(f'events_{yr}_{mo}', None)
                except Exception:
                    pass
            return True
    return False


@app.route('/calendar', methods=['GET', 'POST'])
def calendar_view():
    today = date.today()
    year  = int(request.args.get('year',  today.year))
    month = int(request.args.get('month', today.month))
    error = ''
    saved = False

    if request.method == 'POST':
        if request.form.get('pw') == MENU_PW:
            action = request.form.get('action', '')
            if action == 'add':
                ds      = request.form.get('ds', '').strip()
                ev_type = request.form.get('type', '행사').strip()
                content = request.form.get('content', '').strip()
                note    = request.form.get('note', '').strip()
                if ds and content:
                    try:
                        save_company_event(ds, ev_type, content, note)
                        saved = True
                    except Exception as e:
                        error = str(e)
                else:
                    error = '날짜와 내용을 입력해주세요.'
            elif action == 'delete':
                ds      = request.form.get('ds', '').strip()
                content = request.form.get('content', '').strip()
                try:
                    delete_company_event(ds, content)
                    saved = True
                except Exception as e:
                    error = str(e)
        else:
            error = '비밀번호가 틀렸습니다.'

    events    = get_company_events(year, month)
    ev_by_day = defaultdict(list)
    for ev in events:
        try:
            day = int(ev['ds'].split('-')[2])
            ev_by_day[day].append(ev)
        except Exception:
            pass

    import calendar as _cal
    first_wd, days_in_month = _cal.monthrange(year, month)
    prev_m = month - 1 if month > 1 else 12
    prev_y = year if month > 1 else year - 1
    next_m = month + 1 if month < 12 else 1
    next_y = year if month < 12 else year + 1

    # 달력 주 구성 (일~토 순서, 일요일 시작)
    # first_wd: Mon=0 ... Sun=6  →  offset: Sun=0, so offset=(first_wd+1)%7
    offset = (first_wd + 1) % 7
    weeks  = []
    week   = [None] * offset
    for d in range(1, days_in_month + 1):
        week.append(d)
        if len(week) == 7:
            weeks.append(week)
            week = []
    if week:
        weeks.append(week + [None] * (7 - len(week)))

    ev_types = ['공휴일', '회의', '행사', '생산', '점검', '기타']
    ev_colors = {
        '공휴일': '#FEE2E2', '회의': '#DBEAFE', '행사': '#D1FAE5',
        '생산': '#FFF3E0', '점검': '#EDE9FE', '기타': '#F3F4F6',
    }

    return render_template('calendar.html',
        year=year, month=month,
        weeks=weeks, ev_by_day=dict(ev_by_day),
        ev_colors=ev_colors, ev_types=ev_types,
        today=today, prev_y=prev_y, prev_m=prev_m,
        next_y=next_y, next_m=next_m,
        error=error, saved=saved,
        updated=datetime.now().strftime('%H:%M'),
    )


@app.route('/api/employees')
def api_employees():
    emps = get_employees()
    result = []
    for e in emps:
        eid = str(e.get('사원번호', '')).strip()
        result.append({
            'emp_id':  eid,
            'name':    e.get('성명', ''),
            'factory': 'MJ 1공장' if eid.upper().startswith('M') else 'SCS 2공장',
        })
    return jsonify(result)


@app.route('/api/refresh')
def api_refresh():
    _clear_cache()
    return jsonify({'ok': True, 'ts': datetime.now().strftime('%H:%M:%S')})


# ── 얼굴인식 ─────────────────────────────────────────────────────
FACE_SHEET = '얼굴인식'

def _ensure_face_sheet(sh):
    try:
        return sh.worksheet(FACE_SHEET)
    except WorksheetNotFound:
        ws = sh.add_worksheet(title=FACE_SHEET, rows=200, cols=5)
        ws.append_row(['사원번호', '성명', '공장', 'descriptor', '등록일시'])
        return ws

def get_face_descriptors():
    def _f():
        try:
            sh = _open_sh()
            ws = _ensure_face_sheet(sh)
            rows = ws.get_all_values()[1:]
            result = []
            for r in rows:
                if len(r) < 4 or not r[0].strip():
                    continue
                try:
                    desc = json.loads(r[3])
                    if isinstance(desc, list) and len(desc) == 128:
                        result.append({
                            'emp_id':     r[0].strip(),
                            'name':       r[1].strip(),
                            'factory':    r[2].strip() if len(r) > 2 else '',
                            'descriptor': desc,
                        })
                except Exception:
                    pass
            return result
        except Exception:
            return []
    return _cached('face_desc', _f, ttl=300)

def save_face_descriptor(emp_id, name, factory, descriptor):
    sh  = _open_sh()
    ws  = _ensure_face_sheet(sh)
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    desc_json = json.dumps(descriptor)
    rows = ws.get_all_values()
    for i, r in enumerate(rows[1:], start=2):
        if r and r[0].strip() == emp_id:
            ws.update(range_name=f'A{i}:E{i}',
                      values=[[emp_id, name, factory, desc_json, now_s]])
            with _lock:
                _cache.pop('face_desc', None)
            return
    ws.append_row([emp_id, name, factory, desc_json, now_s],
                  value_input_option='USER_ENTERED')
    with _lock:
        _cache.pop('face_desc', None)


@app.route('/checkin')
def checkin():
    return render_template('checkin.html')


@app.route('/register_face')
def register_face():
    return render_template('register_face.html')


@app.route('/api/face_descriptors')
def api_face_descriptors():
    return jsonify(get_face_descriptors())


@app.route('/api/register_face_frame', methods=['POST'])
def api_register_face_frame():
    """모바일 얼굴 등록 — 단계별 이미지 저장 (face_pending 테이블)."""
    import base64 as _b64
    body   = request.get_json(silent=True) or {}
    emp_id = body.get('emp_id', '').strip().upper()
    phase  = body.get('phase', 0)
    img_b64 = body.get('image', '')
    pw     = body.get('pw', '')

    if not emp_id or not img_b64:
        return jsonify({'ok': False, 'msg': '필수 값 누락'})
    if pw != ADMIN_PW:
        return jsonify({'ok': False, 'msg': '관리자 비밀번호 오류'})

    try:
        img_bytes = _b64.b64decode(img_b64.split(',')[-1])
    except Exception:
        return jsonify({'ok': False, 'msg': '이미지 디코딩 실패'})

    try:
        import db_pg
        if db_pg.is_available():
            with db_pg.cursor() as cur:
                cur.execute(
                    "INSERT INTO face_pending (emp_id, phase, image_data) VALUES (%s, %s, %s)",
                    (emp_id, int(phase), img_bytes))
            return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)[:80]})


@app.route('/api/meal_checkin', methods=['POST'])
def api_meal_checkin():
    body   = request.get_json(silent=True) or {}
    emp_id = body.get('emp_id', '').strip()
    action = body.get('action', '').strip()

    if not emp_id or action not in ('중식신청', '저녁도시락'):
        return jsonify({'ok': False, 'error': '잘못된 요청'})

    settings = get_op_settings()
    now      = _now_kst()
    if action == '중식신청':
        deadline = settings.get('중식신청마감', _OP_DEFAULTS['중식신청마감'])
        if now > deadline:
            return jsonify({'ok': False, 'error': f'중식 사전 신청 마감 시간이 지났습니다. (마감 {deadline})'})
    elif action == '저녁도시락':
        deadline = settings.get('저녁마감', _OP_DEFAULTS['저녁마감'])
        if now > deadline:
            return jsonify({'ok': False, 'error': f'저녁도시락 신청 마감 시간이 지났습니다. (마감 {deadline})'})

    ds    = date.today().strftime('%Y-%m-%d')
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    emps     = get_employees()
    emp      = next((e for e in emps if str(e.get('사원번호', '')).strip() == emp_id), {})
    emp_name = emp.get('성명', '')
    dept     = emp.get('부서명', '')
    rank     = emp.get('직급', '')
    gender   = emp.get('성별', '')
    factory  = 'MJ 1공장' if emp_id.upper().startswith('M') else 'SCS 2공장'

    try:
        sh = _open_sh()
        if action == '중식신청':
            ws   = sh.worksheet('중식신청')
            rows = ws.get_all_values()[1:]
            for r in rows:
                if len(r) >= 4 and r[0].strip() == ds and r[3].strip() == emp_id:
                    return jsonify({'ok': False, 'error': '이미 신청됨'})
            ws.append_row([ds, now_s, factory, emp_id, emp_name, dept, '신청'],
                          value_input_option='USER_ENTERED')
        else:
            ws   = sh.worksheet('저녁도시락')
            rows = ws.get_all_values()[1:]
            for r in rows:
                if len(r) >= 3 and r[0].strip() == ds and r[2].strip() == emp_id:
                    return jsonify({'ok': False, 'error': '이미 신청됨'})
            ws.append_row([ds, now_s, emp_id, emp_name, dept, rank, gender, '웹신청'],
                          value_input_option='USER_ENTERED')

        with _lock:
            _cache.pop(f'meal_{ds}', None)
        return jsonify({'ok': True, 'name': emp_name, 'dept': dept})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/save_face_descriptor', methods=['POST'])
def api_save_face_descriptor():
    body       = request.get_json(silent=True) or {}
    if body.get('pw', '') != MENU_PW:
        return jsonify({'ok': False, 'error': '비밀번호 오류'})
    emp_id     = body.get('emp_id', '').strip()
    name       = body.get('name', '').strip()
    factory    = body.get('factory', '').strip()
    descriptor = body.get('descriptor', [])
    if not emp_id or not isinstance(descriptor, list) or len(descriptor) != 128:
        return jsonify({'ok': False, 'error': '잘못된 데이터'})
    try:
        save_face_descriptor(emp_id, name, factory, descriptor)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ── 당직자 ────────────────────────────────────────────────────────
def get_duty_officers(year: int = None, month: int = None) -> list:
    """당직자 시트에서 해당 연월의 배정 목록 반환. year/month 없으면 전체."""
    def _f():
        try:
            sh = _open_sh()
            ws = sh.worksheet(DUTY_SHEET)
            rows = ws.get_all_records()
            return [{'date': str(r.get('날짜', '')).strip(),
                     'weekday': str(r.get('요일', '')).strip(),
                     'name': str(r.get('당직자', '')).strip()}
                    for r in rows if str(r.get('날짜', '')).strip()]
        except Exception:
            return []
    key = 'duty_all'
    data = _cached(key, _f, ttl=300)
    if year and month:
        prefix = f'{year}-{month:02d}-'
        return [r for r in data if r['date'].startswith(prefix)]
    return data


def get_duty_for_date(ds: str) -> str:
    """특정 날짜(YYYY-MM-DD)의 당직자 이름. 없으면 빈 문자열."""
    all_data = _cached('duty_all', lambda: get_duty_officers(), ttl=300)
    for r in all_data:
        if r['date'] == ds:
            return r['name']
    return ''


def get_weekend_duty_near(ds: str) -> dict:
    """ds 날짜 기준 가장 가까운 주말(토/일)의 당직자 반환."""
    d = datetime.strptime(ds, '%Y-%m-%d').date()
    # 오늘이 토/일이면 자신, 아니면 이번 주 토요일과 다음 일요일 탐색
    result = {}
    for delta in range(-7, 15):
        check = d + timedelta(days=delta)
        wd = check.weekday()
        if wd in (5, 6):
            cs = check.strftime('%Y-%m-%d')
            name = get_duty_for_date(cs)
            label = '토' if wd == 5 else '일'
            key = f'{cs} ({label})'
            result[key] = name or '미배정'
        if len(result) >= 4:
            break
    return result


# ── 근태 신청 (모바일) ────────────────────────────────────────────
@app.route('/attendance_request')
def attendance_request():
    ds     = date.today().strftime('%Y-%m-%d')
    emps   = get_employees()
    op     = get_op_settings()
    # 이번 주말 당직자 조회 (특근 신청서용)
    duty   = get_weekend_duty_near(ds)
    menu_type = op.get('식당유형', '식당')  # '식당' or '중국집'
    return render_template('attendance_request.html',
                           today_ds=ds,
                           employees=emps,
                           duty=duty,
                           menu_type=menu_type,
                           op=op)


@app.route('/api/attendance_request', methods=['POST'])
def api_attendance_request():
    body     = request.get_json(silent=True) or {}
    emp_id   = body.get('emp_id', '').strip()
    ds       = body.get('date', '').strip()
    att_type = body.get('att_type', '').strip()
    value    = body.get('value', '')
    memo     = body.get('memo', '').strip()
    approver = body.get('approver', '').strip()  # 주 52H/64H 초과 시 승인자 성명

    if not emp_id or not ds or not att_type:
        return jsonify({'ok': False, 'error': '필수 항목 누락'})

    try:
        datetime.strptime(ds, '%Y-%m-%d')
    except ValueError:
        return jsonify({'ok': False, 'error': '날짜 형식 오류'})

    emps = get_employees()
    emp  = next((e for e in emps if str(e.get('사원번호', '')).strip() == emp_id), None)
    if not emp:
        return jsonify({'ok': False, 'error': '사원 미등록'})

    # ── 주 52H/64H 경고 (잔업/특근만) ────────────────────────
    if att_type in ATT_CAT_OT:
        try:
            val_f = float(value)
        except Exception:
            val_f = 0.0
        year  = int(ds[:4]); month = int(ds[5:7])
        day   = int(ds[8:10])
        fw    = date(year, month, 1).weekday()
        target_wk = (day + fw - 1) // 7 + 1
        att_rows = get_att_records(year, month)
        cur_ot = 0.0; cur_lv_h = 0.0
        for r in att_rows:
            if str(r.get('사원번호','')).strip() != emp_id:
                continue
            try: d2 = int(str(r.get('일자','')).split('-')[-1])
            except: continue
            wk = (d2 + fw - 1) // 7 + 1
            if wk != target_wk: continue
            atype2 = str(r.get('근태유형','')).strip()
            try: v = float(r.get('값', 0) or 0)
            except: v = 0.0
            if atype2 in ATT_CAT_OT:    cur_ot   += v
            elif atype2 in ATT_CAT_LEAVE: cur_lv_h += v * 8
        new_ot  = cur_ot + val_f
        avail52 = OT_DANGER_H + cur_lv_h
        avail64 = OT_MAX64_H  + cur_lv_h
        if new_ot >= avail64:
            limit_h = 64; required = True
        elif new_ot >= avail52:
            limit_h = 52; required = True
        else:
            limit_h = 0;  required = False
        if required and not approver:
            return jsonify({
                'ok': False,
                'need_approval': True,
                'limit_h': limit_h,
                'new_ot': round(new_ot, 1),
                'avail_h': int(avail64 if limit_h == 64 else avail52),
                'wk_no': target_wk,
                'error': f'주 {limit_h}시간 한도 초과 — 상위 관리자 승인이 필요합니다.'
            })
        if required and approver:
            memo = (memo + ' ' if memo else '') + f'[승인:{approver}/{limit_h}H초과]'

    try:
        year  = int(ds[:4])
        month = int(ds[5:7])
        now_s = datetime.now(KST).strftime('%Y-%m-%d %H:%M:%S')
        sh    = _open_sh()
        ws    = sh.worksheet('근태기록')
        ws.append_row([
            year, month, ds,
            'MJ 1공장' if emp_id.upper().startswith('M') else 'SCS 2공장',
            emp.get('부서명', ''),
            emp_id, emp.get('성명', ''), emp.get('직급', ''),
            att_type, value, memo, '모바일신청', now_s
        ], value_input_option='USER_ENTERED')
        with _lock:
            _cache.pop(f'att_{year}_{month}', None)
        return jsonify({'ok': True, 'name': emp.get('성명', '')})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ── 외부손님 신청 (모바일) ────────────────────────────────────────
@app.route('/guest_request')
def guest_request():
    ds   = date.today().strftime('%Y-%m-%d')
    emps = get_employees()
    return render_template('guest_request.html',
                           today_ds=ds, employees=emps)


@app.route('/api/guest_request', methods=['POST'])
def api_guest_request():
    body      = request.get_json(silent=True) or {}
    mgr_id    = body.get('mgr_id', '').strip()
    visit_dt  = body.get('visit_date', '').strip()
    company   = body.get('company', '').strip()
    gname     = body.get('guest_name', '').strip()
    reason    = body.get('reason', '').strip()
    try: cnt  = max(1, int(body.get('count', 1)))
    except: cnt = 1

    if not mgr_id or not visit_dt or not company or not gname:
        return jsonify({'ok': False, 'error': '담당자/방문일/회사/손님 성명 필수'})
    try:
        datetime.strptime(visit_dt, '%Y-%m-%d')
    except ValueError:
        return jsonify({'ok': False, 'error': '방문예정일 형식 오류'})

    emps = get_employees()
    mgr  = next((e for e in emps if str(e.get('사원번호','')).strip() == mgr_id), None)
    if not mgr:
        return jsonify({'ok': False, 'error': f'담당자 미등록: {mgr_id}'})

    now_d = date.today().strftime('%Y-%m-%d')
    now_t = datetime.now(KST).strftime('%H:%M:%S')
    row   = [visit_dt, now_d, now_t,
             mgr_id, mgr.get('성명',''), mgr.get('부서명',''),
             company, gname, reason, str(cnt)]
    try:
        sh = _open_sh()
        sh.worksheet('외부손님').append_row(row, value_input_option='RAW')
        _guest_cache.append(row)
        # 캐시 무효화
        with _lock:
            for k in list(_cache.keys()):
                if k.startswith('meal_'): _cache.pop(k, None)
        return jsonify({'ok': True,
                        'msg': f'{gname} ({company}) {cnt}명 등록 완료 — 담당:{mgr.get("성명","")}'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/duty')
def api_duty():
    """당직자 목록 API."""
    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)
    return jsonify(get_duty_officers(year, month))


# ═══════════════════════════════════════════════════════════════
# v1.2: 관리자 통합 대시보드 (설비 PM + 개선제안)
# ═══════════════════════════════════════════════════════════════
import db_pg

@app.route('/equipment')
def equipment_dashboard():
    """설비 PM 대시보드 — 관리자 (그룹장/공장장/임원/사장) 공통."""
    if not db_pg.is_available():
        return render_template('error.html',
            error='Supabase 환경변수(SUPABASE_DB_URL) 미설정.\n'
                  'Render 대시보드 → Environment 에 .env 값을 등록하세요.')
    try:
        # KPI
        kpi = {
            'total':  db_pg.query_one("SELECT COUNT(*) c FROM eq_machines WHERE active=TRUE")['c'],
            'issues': db_pg.query_one("SELECT COUNT(*) c FROM eq_issues WHERE status IN ('신규','이관','점검중')")['c'],
            'today':  db_pg.query_one("SELECT COUNT(*) c FROM eq_issues WHERE DATE(occurred_at)=CURRENT_DATE")['c'],
            'this_m': db_pg.query_one(
                "SELECT COUNT(*) c FROM eq_issues WHERE status='완료' "
                "AND DATE_TRUNC('month', closed_at)=DATE_TRUNC('month', CURRENT_DATE)")['c'],
        }
        # 진행중 이슈 (최근 30)
        issues = db_pg.query("""
            SELECT issue_id, occurred_at, eq_id, process_id, major_type, detail, status
            FROM eq_issues WHERE status IN ('신규','이관','점검중')
            ORDER BY occurred_at DESC LIMIT 30
        """)
        # 공정별 이슈 빈도 (이번 달)
        by_process = db_pg.query("""
            SELECT process_id, COUNT(*) c FROM eq_issues
            WHERE DATE_TRUNC('month', occurred_at) = DATE_TRUNC('month', CURRENT_DATE)
            GROUP BY process_id ORDER BY c DESC LIMIT 10
        """)
        return render_template('equipment.html',
                               kpi=kpi, issues=issues, by_process=by_process,
                               updated=datetime.now().strftime('%H:%M'))
    except Exception as e:
        return render_template('error.html', error=f'설비 대시보드: {e}')


@app.route('/equipment/search')
def equipment_search():
    """설비 이력 검색 (모바일) — 누적 이슈 검색만 가능. 접수는 데스크탑.
    현장에서 동일 문제 사례 빠르게 확인용."""
    if not db_pg.is_available():
        return render_template('error.html',
            error='Supabase 환경변수(SUPABASE_DB_URL) 미설정.')
    try:
        kw = request.args.get('q', '').strip()
        eq = request.args.get('eq', '').strip()
        results = []
        if kw or eq:
            sql = """
                SELECT issue_id, occurred_at, eq_id, process_id,
                       major_type, minor_type, detail, cause, action_taken,
                       status, source_tag
                FROM eq_issues WHERE 1=1
            """
            params = []
            if eq:
                sql += " AND eq_id = %s"; params.append(eq)
            if kw:
                sql += """ AND (
                    to_tsvector('simple',
                        COALESCE(detail,'') || ' ' || COALESCE(cause,'') || ' ' ||
                        COALESCE(action_taken,'') || ' ' || COALESCE(major_type,'') || ' ' ||
                        COALESCE(minor_type,''))
                    @@ plainto_tsquery('simple', %s)
                    OR detail ILIKE %s
                    OR action_taken ILIKE %s
                )"""
                params.extend([kw, f'%{kw}%', f'%{kw}%'])
            sql += " ORDER BY occurred_at DESC LIMIT 100"
            results = db_pg.query(sql, tuple(params))
        # 설비 목록 (선택용)
        machines = db_pg.query(
            "SELECT eq_id FROM eq_machines WHERE active=TRUE ORDER BY eq_id")
        return render_template('equipment_search.html',
                               kw=kw, sel_eq=eq, results=results,
                               machines=machines,
                               updated=datetime.now().strftime('%H:%M'))
    except Exception as e:
        return render_template('error.html', error=f'검색: {e}')


@app.route('/improvement')
def improvement_dashboard():
    """개선제안 대시보드 — 5년치 검색 + 통계."""
    if not db_pg.is_available():
        return render_template('error.html',
            error='Supabase 환경변수(SUPABASE_DB_URL) 미설정.')
    try:
        kw = request.args.get('q', '').strip()
        kpi = {
            'total': db_pg.query_one("SELECT COUNT(*) c FROM imp_suggestions")['c'],
            'this_y': db_pg.query_one(
                "SELECT COUNT(*) c FROM imp_suggestions WHERE year=EXTRACT(YEAR FROM CURRENT_DATE)")['c'],
            'adopt': db_pg.query_one(
                "SELECT COUNT(*) c FROM imp_suggestions WHERE status='채택'")['c'],
            'wait': db_pg.query_one(
                "SELECT COUNT(*) c FROM imp_sustainability WHERE status='대기'")['c'],
        }
        years = db_pg.query(
            "SELECT year, COUNT(*) c FROM imp_suggestions "
            "WHERE year IS NOT NULL GROUP BY year ORDER BY year")
        # 연도×분기 누적 데이터
        qstats = db_pg.query(
            "SELECT year, quarter, COUNT(*) c FROM imp_suggestions "
            "WHERE year IS NOT NULL GROUP BY year, quarter ORDER BY year")
        if kw:
            results = db_pg.query("""
                SELECT year, quarter, receipt_no, proposer_name, title, grade
                FROM imp_suggestions
                WHERE to_tsvector('simple', COALESCE(title,'') || ' ' ||
                                            COALESCE(current_state,'') || ' ' ||
                                            COALESCE(idea,'')) @@ plainto_tsquery('simple', %s)
                   OR title ILIKE %s
                ORDER BY year DESC LIMIT 50
            """, (kw, f'%{kw}%'))
        else:
            results = db_pg.query("""
                SELECT year, quarter, receipt_no, proposer_name, title, grade
                FROM imp_suggestions
                ORDER BY year DESC, sug_id DESC LIMIT 30
            """)
        top = db_pg.query(
            "SELECT proposer_name, COUNT(*) c FROM imp_suggestions "
            "WHERE proposer_name IS NOT NULL "
            "GROUP BY proposer_name ORDER BY c DESC LIMIT 10")
        return render_template('improvement.html',
                               kpi=kpi, years=years, qstats=qstats,
                               results=results, top=top,
                               kw=kw, updated=datetime.now().strftime('%H:%M'))
    except Exception as e:
        return render_template('error.html', error=f'개선제안 대시보드: {e}')


def _parse_award_file(wb, fname):
    """업로드된 워크북 → award_history 행 리스트.
    3가지 양식 자동 인식:
      A. 분기 포상 ('X분기 포상' 시트) → quarterly_base
      B. 연말 최대점수 ('누적합계' 시트, '_10만 포인트') → annual_top_score
      C. 연말 최우수 ('최우수' 시트, '_50만 포인트') → annual_top_impact
    """
    import re as _re
    rows = []   # (year, quarter, kind, name, prop_cnt, drv_cnt, score, points, sug_id, receipt_no, title)

    # 파일명에서 연도 추출 (예: "25년..." → 2025)
    m_yr = _re.search(r'(\d{2,4})\s*년', fname)
    yr_from_name = None
    if m_yr:
        yy = int(m_yr.group(1))
        yr_from_name = yy if yy >= 1000 else (2000 + yy)

    for sn in wb.sheetnames:
        ws = wb[sn]
        # 시트 제목에서 연도/분기 힌트
        title_cell = str(ws.cell(1, 1).value or '') + ' ' + str(ws.cell(1, 2).value or '')
        m_y2 = _re.search(r'(\d{4})\s*년', title_cell)
        year = int(m_y2.group(1)) if m_y2 else yr_from_name
        m_q = _re.search(r'([1-4])\s*분기', sn) or _re.search(r'([1-4])\s*분기', title_cell)
        quarter = int(m_q.group(1)) if m_q else None

        # 양식 A: 분기 기본 포상
        if '분기' in sn and '포상' in sn:
            if not year or not quarter: continue
            # R4 헤더: no | 구분(이름) | 제안건수 | 추진건수 | 점수 | 금액
            for ri in range(5, ws.max_row + 1):
                no_ = ws.cell(ri, 1).value
                name = ws.cell(ri, 2).value
                if not no_ or not name: continue
                prop_c = _to_int(ws.cell(ri, 3).value)
                drv_c  = _to_int(ws.cell(ri, 4).value)
                score  = _to_int(ws.cell(ri, 5).value)
                amount = _to_int(ws.cell(ri, 6).value)
                if not amount: continue
                rows.append((year, quarter, 'quarterly_base', str(name).strip(),
                             prop_c, drv_c, score, amount, None, None, None))

        # 양식 B: 연말 최대 점수 누적합계
        elif sn == '누적합계' or '누계' in title_cell or '누적' in title_cell:
            if not year: continue
            # 1위만 100,000P (또는 모두 기록 후 표시 시 1위만)
            for ri in range(4, ws.max_row + 1):
                name = ws.cell(ri, 2).value
                if not name: continue
                prop_c = _to_int(ws.cell(ri, 3).value)
                drv_c  = _to_int(ws.cell(ri, 4).value)
                score  = _to_int(ws.cell(ri, 5).value)
                if score is None: continue
                # 1위만 포상 등록
                pts = 100000 if (ri == 4) else 0
                if pts == 0: continue
                rows.append((year, None, 'annual_top_score', str(name).strip(),
                             prop_c, drv_c, score, pts, None, None, None))

        # 양식 C: 연말 최우수 (개선제안 평가서 형식)
        elif sn == '최우수' or '최우수' in fname:
            if not year: continue
            # R1: 접수no | 부서 | 제안자
            receipt = str(ws.cell(2, 2).value or '').strip()
            proposer = str(ws.cell(2, 8).value or '').strip()
            title    = str(ws.cell(4, 2).value or '').strip()
            driver   = str(ws.cell(4, 8).value or '').strip()
            score    = _to_int(ws.cell(4, 10).value)
            # 제안자 + 추진자 둘 다 500,000P 각각? 또는 묶음? — 일단 둘 다 등록
            for awardee in [proposer, driver]:
                if not awardee or awardee in ('-', '없음'): continue
                rows.append((year, None, 'annual_top_impact', awardee.strip(),
                             None, None, score, 500000, None, receipt, title))
    return rows


def _to_int(v):
    if v is None or v == '': return None
    try:
        s = str(v).replace(',', '').replace('P', '').strip()
        return int(float(s))
    except (ValueError, TypeError):
        return None


@app.route('/improvement/upload_award', methods=['GET', 'POST'])
def improvement_upload_award():
    """분기/연말 포상 확정 파일 업로드 → DB 자동 반영."""
    if not db_pg.is_available():
        return render_template('error.html',
            error='Supabase 환경변수(SUPABASE_DB_URL) 미설정.')
    msg, ok = None, None
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            msg = '파일을 선택하세요.'
        else:
            try:
                from openpyxl import load_workbook
                import io as _io
                wb = load_workbook(_io.BytesIO(f.read()), read_only=False, data_only=True)
                parsed = _parse_award_file(wb, f.filename)
                wb.close()
                # 사번 자동 매칭 + UPSERT
                inserted = 0
                for (year, q, kind, name, pc, dc, sc, pts, sid, rno, ttl) in parsed:
                    # 이름 → 사번 (있으면)
                    emp_id = None
                    r = db_pg.query_one(
                        "SELECT emp_id FROM employees "
                        "WHERE REPLACE(name, ' ', '') = REPLACE(%s, ' ', '') LIMIT 1",
                        (name,))
                    if r: emp_id = r['emp_id']
                    with db_pg.cursor() as cur:
                        cur.execute("""
                            INSERT INTO imp_award_history (
                                year, quarter, kind, awardee_name, awardee_id,
                                proposal_count, driver_count, score, points,
                                sug_id, receipt_no, title, source_file
                            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            ON CONFLICT (year, quarter, kind, awardee_name) DO UPDATE SET
                                proposal_count = EXCLUDED.proposal_count,
                                driver_count = EXCLUDED.driver_count,
                                score = EXCLUDED.score,
                                points = EXCLUDED.points,
                                awardee_id = COALESCE(EXCLUDED.awardee_id, imp_award_history.awardee_id),
                                receipt_no = COALESCE(EXCLUDED.receipt_no, imp_award_history.receipt_no),
                                title = COALESCE(EXCLUDED.title, imp_award_history.title),
                                source_file = EXCLUDED.source_file,
                                uploaded_at = NOW()
                        """, (year, q, kind, name, emp_id, pc, dc, sc, pts,
                              sid, rno, ttl, f.filename))
                    inserted += 1
                ok = True
                msg = f'✓ 업로드 완료\n{inserted}건 DB 자동 반영됨\n파일: {f.filename}'
            except Exception as e:
                msg = f'⚠ 분석 오류: {e}'
    return render_template('improvement_upload.html', msg=msg, ok=ok)


@app.route('/improvement/stats')
def improvement_stats():
    """개선제안 통계 — 연도/분기별 + 포상 (1점=1000P)."""
    if not db_pg.is_available():
        return render_template('error.html',
            error='Supabase 환경변수(SUPABASE_DB_URL) 미설정.')
    try:
        year = request.args.get('year', type=int)
        quarter = request.args.get('quarter', type=int)

        period_stats = db_pg.query("""
            SELECT year, quarter,
                   COUNT(*) AS total,
                   COUNT(CASE WHEN grade IN ('S+','S','A') THEN 1 END) AS high_grade,
                   SUM(COALESCE(reward_points, 0)) AS reward_sum
            FROM imp_suggestions
            WHERE year IS NOT NULL
            GROUP BY year, quarter
            ORDER BY year DESC, quarter DESC
        """)
        proposer_stats = []
        if year:
            sql = ("SELECT proposer_name, COUNT(*) AS total, "
                   "SUM(COALESCE(reward_points, 0)) AS reward_sum, "
                   "STRING_AGG(DISTINCT grade, ',' ORDER BY grade) AS grades "
                   "FROM imp_suggestions WHERE year = %s "
                   "AND proposer_name IS NOT NULL")
            params = [year]
            if quarter:
                sql += " AND quarter = %s"; params.append(quarter)
            sql += " GROUP BY proposer_name ORDER BY total DESC"
            proposer_stats = db_pg.query(sql, tuple(params))

        # 확정 포상 이력 (파일 업로드로 등록된 것 우선)
        # 분기 기본 포상
        quarterly_base = db_pg.query(
            "SELECT * FROM imp_award_history WHERE kind='quarterly_base' "
            "ORDER BY year DESC, quarter DESC, points DESC")
        # 분기 최다 (확정본 우선, 없으면 동적 계산 view)
        quarterly_top = db_pg.query(
            "SELECT * FROM imp_award_history "
            "WHERE kind IN ('quarterly_top_propose', 'quarterly_top_drive') "
            "ORDER BY year DESC, quarter DESC, kind")
        if not quarterly_top:
            # 확정 전이면 자동 계산 표시 (참고용)
            quarterly_top = db_pg.query(
                "SELECT year, quarter, "
                "       CASE kind WHEN 'top_proposer' THEN 'quarterly_top_propose' "
                "                 ELSE 'quarterly_top_drive' END AS kind, "
                "       name AS awardee_name, cnt, points "
                "FROM v_quarterly_award "
                "ORDER BY year DESC, quarter DESC, kind")
        # 연말 최대점수
        annual_top_score = db_pg.query(
            "SELECT * FROM imp_award_history WHERE kind='annual_top_score' "
            "ORDER BY year DESC")
        # 연말 최우수
        annual_top_impact = db_pg.query(
            "SELECT * FROM imp_award_history WHERE kind='annual_top_impact' "
            "ORDER BY year DESC")

        years = sorted({r['year'] for r in period_stats}, reverse=True)
        return render_template('improvement_stats.html',
                               period_stats=period_stats,
                               proposer_stats=proposer_stats,
                               quarterly_base=quarterly_base,
                               quarterly_awards=quarterly_top,
                               annual_top_score=annual_top_score,
                               annual_top_impact=annual_top_impact,
                               years=years, sel_year=year, sel_quarter=quarter,
                               updated=datetime.now().strftime('%H:%M'))
    except Exception as e:
        return render_template('error.html', error=f'통계: {e}')


@app.route('/set_tz/<tz>')
def set_tz(tz):
    """사용자 시간대 변경 (쿠키 1년)."""
    if tz not in TZ_MAP:
        return jsonify({'ok': False, 'error': 'unknown tz'}), 400
    nxt = request.args.get('next') or '/'
    resp = redirect(nxt)
    resp.set_cookie('tz', tz, max_age=365*24*3600, samesite='Lax')
    return resp


@app.context_processor
def _inject_version():
    """모든 템플릿에 버전 + 시간대 정보 주입."""
    tz = user_tz()
    tz_name = ('PHT' if tz == PHT else 'KST')
    return {
        'mjt_version': '1.3.0',
        'mjt_version_date': '2026-05-31',
        'user_tz': tz_name,
        'fmt_local': fmt_local,
    }


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
